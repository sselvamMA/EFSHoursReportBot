import pandas as pd
from openpyxl import load_workbook
import openpyxl
import streamlit as st
from tempfile import NamedTemporaryFile
import tempfile
import os
from pathlib import Path
from fastai.vision import *
from fastai import *
import pickle


def csv_extraction(file):
    new_file = pd.read_csv(file)

st.title('EFS Report Bot')

csv_name = ''
excel_file_name = ''

week_in_question = st.radio(
    "What week of the pay period are you inputting data from?",
    ('Week 1 of the Pay Period', 'Week 2 of the Pay Period'))

cell_name = ''

if week_in_question == 'Week 1 of the Pay Period':
    cell_name = 'C%d'
else:
    cell_name = 'F%d'

csv_file = st.file_uploader('Upload EFS Data from the week', type = 'csv')
data = None
if csv_file is not None:
    csv_file.seek(0)
    data = pd.read_csv(csv_file, low_memory=False)

if csv_file and week_in_question:
    
    excel_file = st.file_uploader('Upload EFS Hours Report from the week', type = 'xlsx')
    if excel_file is not None:
        df1 = pd.read_excel(excel_file)
        hours_1 = data
        new_df = hours_1.iloc[:, 7]
        paycode_columns = hours_1.iloc[:, 6]

        hours_worked_per_employee = []
        skip_next = False
        for i in range(len(new_df)):
            if skip_next:
                skip_next = False
                continue
            if(paycode_columns[i] == 'OVTIME'):
                hours_worked_per_employee.append(new_df[i] + new_df[i+1])
                skip_next = True
            else:
                hours_worked_per_employee.append(new_df[i])
        print(hours_worked_per_employee)

        fName_df = pd.DataFrame(hours_1.iloc[:, 2])
        lName_df = pd.DataFrame(hours_1.iloc[:, 1])
        combined_values = (lName_df['Last Name'].map(str)) + ', ' + (fName_df['First Name'].map(str))

        name_list = []
        seen_elements = set()
        for j in range(len(combined_values)):
            if combined_values[j] not in seen_elements:
                name_list.append(combined_values[j])
                seen_elements.add(combined_values[j])

        employee_hours_dict = dict(zip(name_list, hours_worked_per_employee))
        print(employee_hours_dict.values())

        for name in name_list:
            if (name in employee_hours_dict.keys()):
                print(name, employee_hours_dict[name])

#insert filepath here
        wb = openpyxl.load_workbook('insert Empty EFS Hours report file path')
        ws = wb['EFS Hours Report']
        names_EFS_standard = []
        for index in range(len(hours_worked_per_employee)):
            names_EFS_standard.append(ws.cell(index+7, 2).value)
        print(names_EFS_standard)

        df = pd.DataFrame(names_EFS_standard)
        for index, row in df.iterrows():
            cell = cell_name  % (index + 7)
            if (names_EFS_standard[index] in employee_hours_dict.keys()):
                ws[cell] = employee_hours_dict[names_EFS_standard[index]]
            else:
                ws[cell] = 'n/a'

#insert filepath here
        wb.save('insert Empty EFS Hours report file path')
        st.balloons()
